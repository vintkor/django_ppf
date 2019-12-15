from pprint import pprint
from assistant.models import Product, Parameter, Category, Photo
from datetime import datetime
from xml.dom import minidom
from django.conf import settings
from bs4 import BeautifulSoup
import requests
from django.utils.crypto import get_random_string
import decimal
from currency.models import Currency
from openpyxl import load_workbook
from django.db.transaction import atomic
import csv
import os
import xml.etree.ElementTree as ET
import logging
from xlsxwriter import Workbook


class ParseHoroz:
    """
    Парсер сайта https://horozua.com -> https://horozua.com/index.php?route=feed/yandex_yml
    """

    def __init__(self, link=None, filename=None, my_currency_code=None):
        self.products = []
        self._link = link
        self._filename = filename
        self.vendor_name = 'Horoz Electric'
        self._my_currency_code = 'USD'
        if my_currency_code:
            self._set_currency()
        self._set_headers()

    def _set_headers(self):
        self.headers = requests.utils.default_headers()
        self.headers.update(
            {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
            }
        )

    def _set_currency(self):
        try:
            currency = Currency.objects.get(code__iexact=self._my_currency_code)
            self._currency = currency
        except Currency.DoesNotExist:
            raise ValueError('В базе нет валюты с указанным кодом')

    def _get_source(self):
        r = requests.get(self._link, headers=self.headers)
        if r.status_code == 200:
            with open(settings.MEDIA_ROOT+'/horoz_parse.xml', 'wb') as f:
                f.write(r.content)
            self._filename = settings.MEDIA_ROOT+'/horoz_parse.xml'

    def set_products(self):
        if self._link:
            self._get_source()
        tree = ET.parse(self._filename)
        root = tree.getroot()
        for offer in root.iter('offer'):
            product_item = dict()
            params = list()
            images = list()
            for product in offer:
                if product.tag == 'name':
                    name = product.text.replace('&quot;', '"')
                    product_item['title'] = name

                product_item['vendor_name'] = self.vendor_name
                product_item['currency_id'] = self._currency.id

                if product.tag == 'vendorCode':
                    product_item['vendor_id'] = product.text

                if product.tag == 'price':
                    product_item['price'] = decimal.Decimal(product.text)

                if product.tag == 'description':
                    if product.text and len(product.text) > 50:
                        product_item['text'] = product.text
                    else:
                        product_item['text'] = ''

                if product.tag == 'delivery':
                    if product.text == 'true':
                        product_item['availability_prom'] = '+'
                    else:
                        product_item['availability_prom'] = '-'

                if product.tag == 'picture':
                    if product.text:
                        images.append(product.text)

                product_item['images'] = images

                if product.tag == 'param':
                    param_name = product.attrib.get('name')
                    param_value = product.text
                    if param_name and param_value:
                        params.append({
                            'name': param_name,
                            'value': param_value,
                        })

                product_item['params'] = params
            self.products.append(product_item)

    def print_products(self):
        pprint(self.products)
        pprint(len(self.products))

    def _get_products_to_update_and_new_products(self):
        vendor_id = [i.get('vendor_id') for i in self.products]
        products_id_in_db = [i.vendor_id for i in Product.objects.filter(vendor_name__iexact=self.vendor_name, vendor_id__in=vendor_id)]

        new_products = list()
        products_to_update = list()

        for product in self.products:
            if product['vendor_id'] in products_id_in_db:
                products_to_update.append(product)
            else:
                new_products.append(product)

        return new_products, products_to_update

    def add_or_update_products_in_db(self):
        new_products, products_to_update = self._get_products_to_update_and_new_products()
        self.products = []

        try:
            test_category = Category.objects.get(title='TEST CATEGORY')
        except Category.DoesNotExist:
            test_category = Category(title='TEST CATEGORY')
            test_category.save()

        for i in products_to_update:
            products = Product.objects.filter(vendor_name__iexact=self.vendor_name, vendor_id__in=i['vendor_id'])
            for product in products:
                product.price = i['price']
                product.availability_prom = i['availability_prom']
                product.save(update_fields=('price', 'availability_prom',))

        for idx, i in enumerate(new_products):

            product = Product(
                title=i['title'],
                vendor_name=i['vendor_name'],
                currency_id=i['currency_id'],
                vendor_id=i['vendor_id'],
                price=i['price'],
                text=i['text'],
                availability_prom=i['availability_prom'],
                category=test_category,
            )
            product.save()

            bulk_images = list()
            urls = i['images']
            for index, url in enumerate(urls):
                filename = self._make_filename(product.pk)
                self._get_and_save_image(url, filename)
                if index == 0:
                    product.image = filename
                    product.save(update_fields=('image',))
                else:
                    bulk_images.append(Photo(product=product, image=filename))
            Photo.objects.bulk_create(bulk_images)

            bulk_params = list()
            for p in i['params']:
                param = Parameter(
                    product=product,
                    parameter=p['name'],
                    value=p['value'],
                )
                bulk_params.append(param)
            Parameter.objects.bulk_create(bulk_params)
            print('Saved {} products'.format(idx + 1))

    def _get_and_save_image(self, url, filename):
        r = requests.get(url, headers=self.headers)
        if r.status_code == 200:
            with open('media/' + filename, 'wb') as file:
                file.write(r.content)

    @staticmethod
    def _make_filename(pk, ext=None):
        name = get_random_string(25)
        return 'images/{}__{}'.format(pk, name)

    @staticmethod
    def _get_file_ext(url):
        return url.split('.')[-1]


class ExportFeatures:
    """
    Импорт параметров из прома на ассистант
    """

    def __init__(self, filename):
        self._filename = filename
        self.products = []

    def parse_features(self):
        tree = ET.parse(self._filename)
        root = tree.getroot()
        for offer in root.iter('offer'):
            product_item = dict()
            params = list()
            for product in offer:

                if product.tag == 'vendorCode':
                    pf_code = product.text
                    product_item['pf_code'] = pf_code

                if product.tag == 'country_of_origin':
                    param_name = 'Страна'
                    param_value = product.text
                    params.append({
                        'name': param_name,
                        'value': param_value,
                    })

                if product.tag == 'param':
                    param_name = product.attrib.get('name')
                    param_value = product.text
                    param_unit = product.attrib.get('unit')
                    if param_unit:
                        params.append({
                            'name': param_name,
                            'value': param_value + param_unit,
                        })
                    else:
                        params.append({
                            'name': param_name,
                            'value': param_value,
                        })

            product_item['params'] = params
            if len(params) > 2:
                self.products.append(product_item)

    def print_products(self, short=True):
        print('-'*80)
        for index, product in enumerate(self.products):
            if short:
                print(index, '---->', product['pf_code'], len(product['params']))
            else:
                print(index, '---->', product['pf_code'], len(product['params']))
                for item in product['params']:
                    print(item['name'], '--->', item['value'])

    def save_params(self):
        for product in self.products:
            with atomic():
                try:
                    p = Product.objects.prefetch_related('parameter_set').get(code=product['pf_code'])

                    for param in p.parameter_set.all():
                        param.delete()

                    for parameter in product['params']:
                        if parameter['name'] and parameter['value']:
                            param = Parameter(
                                product=p,
                                parameter=parameter['name'],
                                value=parameter['value'],
                            )
                            param.save()
                except Product.DoesNotExist:
                    pass


def clear_content(content):
    soup = BeautifulSoup(content, "html.parser")

    for img in soup('img'):
        img.decompose()

    for a in soup('a'):
        a.decompose()

    for iframe in soup('iframe'):
        iframe.decompose()

    return str(soup)


def make_xml(products=None):
    if products:
        products = products
    else:
        products = Product.objects.filter(import_to_rozetka=True)

    category_list = []
    for product in products:
        if product.import_to_rozetka:
            [
                category_list.append(j) for j in [
                    i for i in product.category_rozetka.get_ancestors(include_self=True, ascending=False)
                ]
            ]

    categories_list = [i for i in set(category_list)]

    imp = minidom.DOMImplementation()
    doctype = imp.createDocumentType(
        qualifiedName="yml_catalog",
        publicId="",
        systemId="shops.dtd",
    )

    doc = minidom.Document()
    doc.appendChild(doctype)

    yml_catalog = doc.createElement('yml_catalog')
    doc.appendChild(yml_catalog)
    yml_catalog.setAttribute('date', datetime.now().strftime("%Y-%m-%d %H:%M"))

    shop = doc.createElement('shop')
    yml_catalog.appendChild(shop)

    name = doc.createElement('name')
    name_text = doc.createTextNode('PPF Company')
    name.appendChild(name_text)
    shop.appendChild(name)

    company = doc.createElement('company')
    company_text = doc.createTextNode('PPF Company inc.')
    company.appendChild(company_text)
    shop.appendChild(company)

    url = doc.createElement('url')
    url_text = doc.createTextNode(settings.SITE_URL)
    url.appendChild(url_text)
    shop.appendChild(url)

    currencies = doc.createElement('currencies')
    shop.appendChild(currencies)

    currency = doc.createElement('currency')
    currency.setAttribute('id', 'UAH')
    currency.setAttribute('rate', '1')
    currencies.appendChild(currency)

    categories = doc.createElement('categories')
    shop.appendChild(categories)

    for cat in categories_list:
        category = doc.createElement('category')
        category_text = doc.createTextNode(cat.title)
        if cat.parent:
            category.setAttribute('parentId', str(cat.parent.id))
        category.setAttribute('id', str(cat.id))
        category.appendChild(category_text)
        categories.appendChild(category)

    offers = doc.createElement('offers')
    shop.appendChild(offers)

    for product in products:
        if product.import_to_rozetka:
            offer = doc.createElement('offer')
            offer.setAttribute('id', '{}'.format(product.id))
            if product.stock_quantity == 0:
                offer.setAttribute('available', 'false')
            else:
                offer.setAttribute('available', 'true')
            offers.appendChild(offer)

            offer_url = doc.createElement('url')
            offer_url_text = doc.createTextNode('{}{}'.format(settings.SITE_URL, product.get_absolute_url()))
            offer_url.appendChild(offer_url_text)
            offer.appendChild(offer_url)

            if product.old_price_percent:
                price = doc.createElement('price')
                price_text = doc.createTextNode(str(product.get_old_price()))
                price.appendChild(price_text)
                offer.appendChild(price)

                old_price = doc.createElement('price_old')
                price_text = doc.createTextNode(str(product.get_price_UAH()))
                old_price.appendChild(price_text)
                offer.appendChild(old_price)
            else:
                price = doc.createElement('price')
                price_text = doc.createTextNode(str(product.get_promo_price()))
                price.appendChild(price_text)
                offer.appendChild(price)

            if product.promo_percent:
                price_promo = doc.createElement('price_promo')
                price_promo_text = doc.createTextNode(str(product.get_promo_price()))
                price_promo.appendChild(price_promo_text)
                offer.appendChild(price_promo)

            currencyId = doc.createElement('currencyId')
            currencyId_text = doc.createTextNode('UAH')
            currencyId.appendChild(currencyId_text)
            offer.appendChild(currencyId)

            categoryId = doc.createElement('categoryId')
            categoryId_text = doc.createTextNode(str(product.category_rozetka.id))  # <--
            categoryId.appendChild(categoryId_text)
            offer.appendChild(categoryId)

            if product.image:
                picture = doc.createElement('picture')
                picture_text = doc.createTextNode('{}{}'.format(settings.SITE_URL, product.image.url))
                picture.appendChild(picture_text)
                offer.appendChild(picture)

            for photo in product.get_images():
                picture_set = doc.createElement('picture')
                picture_set_text = doc.createTextNode('{}{}'.format(settings.SITE_URL, photo.image.url))
                picture_set.appendChild(picture_set_text)
                offer.appendChild(picture_set)

            vendor = doc.createElement('vendor')
            vendor_text = doc.createTextNode(product.manufacturer.title)
            vendor.appendChild(vendor_text)
            offer.appendChild(vendor)

            stock_quantity = doc.createElement('stock_quantity')
            stock_quantity_text = doc.createTextNode(str(product.stock_quantity))
            stock_quantity.appendChild(stock_quantity_text)
            offer.appendChild(stock_quantity)

            name = doc.createElement('name')
            name_text = doc.createTextNode('{title} ({code})'.format(
                title=product.title,
                code=product.code,
            ))
            name.appendChild(name_text)
            offer.appendChild(name)

            description = doc.createElement('description')
            cdata = doc.createCDATASection(clear_content(product.text))
            description.appendChild(cdata)
            offer.appendChild(description)

            # param1 = doc.createElement('param')
            # param1_text = doc.createTextNode(product.code)
            # param1.appendChild(param1_text)
            # param1.setAttribute('name', 'Артикул')
            # offer.appendChild(param1)

            for feature in product.parameter_set.filter(is_dop_param_for_rozetka=False):
                param2 = doc.createElement('param')
                param2_text = doc.createTextNode(feature.value)
                param2.appendChild(param2_text)
                param2.setAttribute('name', feature.parameter)
                offer.appendChild(param2)

            dop_parameters = []
            for dop_param in product.parameter_set.filter(is_dop_param_for_rozetka=True):
                dop_parameters.append('{}: {}<br/>'.format(dop_param.parameter, dop_param.value))

            param3 = doc.createElement('param')
            param3.setAttribute('name', 'Дополнительные характеристики')
            cdata2 = doc.createCDATASection(''.join(dop_parameters))
            param3.appendChild(cdata2)
            offer.appendChild(param3)

    file_handle = open("rozetka.xml", "w")
    doc.writexml(file_handle, encoding='UTF-8')
    file_handle.close()


def make_xlsx_for_prom():
    file_name = 'prom.xlsx'
    queryset = Product.objects.select_related('category', 'unit').filter(active=True, import_to_prom=True)

    workbook = Workbook(file_name)
    worksheet = workbook.add_worksheet('Export Products Sheet')
    worksheet_2 = workbook.add_worksheet('Export Groups Sheet')

    header = (
        'Название_позиции',
        'Ключевые_слова',
        'Описание',
        'Тип_товара',
        'Цена',
        'Валюта',
        'Единица_измерения',
        'Ссылка_изображения',
        'Наличие',
        'Идентификатор_товара',
        'Идентификатор_группы',
        'Код_товара',
        'Номер_группы',
    )

    [worksheet.write(0, col, i) for col, i in enumerate(header)]

    for row, item in enumerate(queryset):
        worksheet.write(row + 1, 0, item.title)
        worksheet.write(row + 1, 1, item.category.title)
        worksheet.write(row + 1, 2, item.text.replace(chr(13), '').replace(chr(10), '').replace(
            'src="/media/uploads/', 'src="{}/media/uploads/'.format(settings.SITE_URL)))
        worksheet.write(row + 1, 3, 'r')
        worksheet.write(row + 1, 4, item.get_price_UAH())
        worksheet.write(row + 1, 5, item.get_currency_code())
        worksheet.write(row + 1, 6, item.get_unit())
        worksheet.write_string(row + 1, 7, ''.join(
            ['{}{}, '.format(settings.SITE_URL, img) for img in item.get_all_photo()]
        ))
        worksheet.write(row + 1, 8, item.availability_prom)
        worksheet.write(row + 1, 9, item.code)
        worksheet.write(row + 1, 10, item.category.id)
        worksheet.write(row + 1, 11, item.code)
        worksheet.write(row + 1, 12, item.category.id)

    header_2 = (
        'Номер_группы',
        'Название_группы',
        'Идентификатор_группы',
        'Номер_родителя',
        'Идентификатор_родителя',
    )

    [worksheet_2.write(0, col, i) for col, i in enumerate(header_2)]

    for row, item in enumerate(Category.objects.all()):
        worksheet_2.write(row + 1, 0, item.id)
        worksheet_2.write(row + 1, 1, item.title)
        worksheet_2.write(row + 1, 2, item.id)
        worksheet_2.write(row + 1, 3, item.get_id())
        worksheet_2.write(row + 1, 4, item.get_id())


def parse_atmosfera():
    file_path = 'http://www.atmosfera.ua/xport2shop/AtmosferaProds2exportUA.xml'
    r = requests.get(file_path)
    soup = BeautifulSoup(r.content, 'xml')

    categories = dict()
    for link in soup.find_all('category'):
        categories[link.get('id')] = {
            'parent_id': link.get('parentId'),
            'title': link.text,
        }

    products = list()
    for product in soup.find_all('item'):
        products.append({
            'id': product.find('Код_товара').text if product.find('Код_товара') else False,
            'price': product.find('Цена').text if product.find('Цена') else False,
            'category': product.find('parentId').text if product.find('parentId') else False,
            'currency': product.find('Валюта').text if product.find('Валюта') else False,
            'unit': product.find('Единица_измерения').text if product.find('Единица_измерения') else False,
            'images': product.find('Ссылка_изображения').text if product.find('Ссылка_изображения') else False,
            'title': product.find('Название_позиции').text if product.find('Название_позиции') else False,
            'vendor': product.find('Производитель').text if product.find('Производитель') else False,
            'country': product.find('Страна_производитель').text if product.find('Страна_производитель') else False,
            'desc': product.find('Описание').text if product.find('Описание') else False,
            'keywords': product.find('Ключевые_слова').text if product.find('Ключевые_слова') else False,
            'params': [{
                'name': p.get('name'),
                'value': p.text,
            } for p in product.find_all('param')],
        })

    print('-------------- CATEGORIES --------------')
    for k, v in categories.items():
        print(k, v)

    print('-------------- PRODUCTS --------------')
    for p in products:
        print(p)


def get_and_save_image(url, filename):
    r = requests.get(url)
    if r.status_code == 200:
        with open('media/' + filename, 'wb') as file:
            file.write(r.content)


def make_filename(pk, ext):
    name = get_random_string(25)
    return 'images/{}__{}'.format(pk, name)


def get_file_ext(url):
    ext = url.split('.')[-1]
    if len(ext) > 3:
        return ''
    return ext


def parse_yantarlk():
    """ Парсер сайта https://yantarlk.com.ua/ """

    try:
        test_category = Category.objects.get(title='TEST CATEGORY')
    except Category.DoesNotExist:
        test_category = Category(title='TEST CATEGORY')
        test_category.save()

    currency = Currency.objects.get(code='UAH')

    url = 'https://yantarlk.com.ua/'
    headers = requests.utils.default_headers()
    headers.update(
        {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
        }
    )
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')

    category_links = [
        i.find('a').get('href') for i in
        soup.find('div', attrs={'id': 'menu2'}).find_all('div', attrs={'class': 'title'})
    ]

    product_links = []
    for category_url in category_links:
        r = requests.get(category_url + '/?limit=2000', headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')
        for i in soup.find_all('h4'):
            link = i.find('a').get('href')
            product_links.append(link)

    # product_links = ['https://yantarlk.com.ua/spec_sredstva_i_material/mastika_oz']
    products = []
    for index, product_link in enumerate(product_links):
        print('---------- Page {}'.format(index))
        r = requests.get(product_link, headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')

        try:
            types_list = soup.find('div', attrs={'class', 'options_no_buy'}).find_all('div', attrs={'class': 'radio'})
        except:
            continue

        params_list = []
        for type in types_list:
            temp_types = type.find_all('span')
            params = dict()
            params['weight'] = temp_types[0].text
            try:
                params['plus_price'] = temp_types[1].text.split('+')[1].split(' ')[0]
            except IndexError:
                pass
            params_list.append(params)

        parameters = []
        for param_row in soup.find_all('tr', attrs={'itemprop': 'additionalProperty'}):
            row = param_row.find_all('td')
            try:
                parameters.append({
                    'name': row[0].text,
                    'value': row[1].text,
                })
            except IndexError:
                pass

        title = soup.find('title').text.split(' - купить ')[0]
        image_link = soup.find('a', attrs={'class': 'main-image'}).get('href')
        price = soup.find('span', attrs={'class': 'update_price'}).text.split('грн')[0]
        desc = soup.find('div', attrs={'id': 'tab-description'}).text

        for pl in params_list:
            prod_dict = {
                'title': title + ' ' + pl['weight'],
                'image': image_link,
                'description': desc,
                'parameters': parameters,
                'weight': pl['weight'],
                'vendor_name': 'ЯНТАРЬ',
            }
            # try:
            #     prod_dict['price'] = int(price) + int(pl.get('plus_price')) if pl.get('plus_price') else price
            # except:
            #     prod_dict['price'] = 0

            products.append(prod_dict)

    print('Finded', len(products), 'products')

    for ind, i in enumerate(products):
        product = Product()
        product.category = test_category
        product.vendor_name = i['vendor_name']
        product.title = i['title']
        product.text = i['description']
        # product.price = decimal.Decimal(i['price'])
        product.stock_quantity = 100
        product.active = True
        product.currency = currency
        product.save()

        parameter = Parameter()
        parameter.product = product
        parameter.parameter = 'Произвадитель'
        parameter.value = i['vendor_name']
        parameter.save()

        parameter2 = Parameter()
        parameter2.product = product
        parameter2.parameter = 'Страна'
        parameter2.value = 'Украина'
        parameter2.save()

        for param in i['parameters']:
            new_param = Parameter()
            new_param.product = product
            new_param.parameter = param['name']
            new_param.value = param['value'][:199]
            new_param.save()

        if i['image']:
            url = i['image']

            ext = get_file_ext(url)
            filename = make_filename(product.pk, ext)
            get_and_save_image(url, filename)

            product.image = filename
            product.save(update_fields=('image',))

        print('Saved', ind + 1, 'from', len(products), 'products')


def parse_mizol():
    file_path = 'http://api.mizol.ua/?modelName=Mizol_Characteristics&calledMethod=getProductCharacteristic&methodProperties[price]=RRC&apiKey=52542&hash=5398382aa014f0a06312ff751c5949c4cbb52f61ba02341dbf67ecedfab68bcf'
    r = requests.get(file_path)
    soup = BeautifulSoup(r.content, 'xml')
    vendor_name = 'Mizol'

    try:
        test_category = Category.objects.get(title='TEST CATEGORY')
    except Category.DoesNotExist:
        test_category = Category(title='TEST CATEGORY')
        test_category.save()

    products_id_in_db = [
        int(product['vendor_id']) for product in Product.objects.filter(vendor_name=vendor_name).values('vendor_id')
    ]

    new_products = list()
    products_for_update = list()

    for ind, product in enumerate(soup.find_all('offer')):
        print('----> finded', ind, 'products')

        if int(product['id']) in products_id_in_db:
            products_for_update.append({
                'vendor_id': int(product['id']),
                'price': decimal.Decimal(product.find('price').text) if product.find('price').text else 0,
            })
        else:
            new_products.append({
                'vendor_id': int(product['id']),
                'vendor_name': vendor_name,
                'title': product.find('name').text,
                'manufacturer': product.find('vendor').text if product.find('vendor').text else None,
                'country': product.find('country').text if product.find('country').text else None,
                'description': product.find('description').text,
                'active': True if product.find('available') == 'true' else False,
                'picture': product.find('picture').text if product.find('picture').text else None,
                'price': decimal.Decimal(product.find('price').text) if product.find('price').text else 0,
            })

    len_new_products = len(new_products)

    currency = Currency.objects.get(code='UAH')

    for ind, i in enumerate(new_products):
        product = Product()
        product.category = test_category
        product.vendor_id = i['vendor_id']
        product.vendor_name = i['vendor_name']
        product.title = i['title']
        product.text = i['description']
        product.active = True
        product.currency = currency
        product.save()

        if i['manufacturer']:
            parameter = Parameter()
            parameter.product = product
            parameter.parameter = 'Произвадитель'
            parameter.value = i['vendor_name']
            parameter.save()

        if i['country']:
            parameter2 = Parameter()
            parameter2.product = product
            parameter2.parameter = 'Страна'
            parameter2.value = i['country']
            parameter2.save()

        if i['picture']:
            url = i['picture']

            ext = get_file_ext(url)
            filename = make_filename(product.pk, ext)
            get_and_save_image(url, filename)

            product.image = filename
            product.save(update_fields=('image',))

        print('----> added', ind, 'items from', len_new_products)


def update_prices(filename, vendor_name):
    file_path = settings.MEDIA_ROOT + '/' + filename
    wb = load_workbook(file_path)
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]

    data_list = []

    for idx, row in enumerate(sheet.rows, start=1):
        if idx > 2 and row[1].value:
            data_list.append({
                'id': row[1].value,
                'available': '-' if row[6].value == 'НЕТ' else '+',
                'stock_quantity_for_rozetka': 0 if row[6].value == 'НЕТ' else 100,
                'price': row[10].value,
            })

    try:
        with atomic():
            for i in data_list:
                products = Product.objects.filter(vendor_id=i['id'], vendor_name=vendor_name)

                for product in products:
                    product.price = decimal.Decimal(i['price'])
                    product.availability_prom = i['available']
                    product.stock_quantity = i['stock_quantity_for_rozetka']
                    product.save(update_fields=('price', 'availability_prom', 'stock_quantity'))
        logging.info('[SUCCESS]: Ubpating product by mizol')
    except:
        logging.info('[ERROR]: Ubpating product by mizol')
    finally:
        os.remove(file_path)


def import_parameters_form_prom(filename):
    import_csv = ExportFeatures(filename)
    import_csv.parse_features()
    import_csv.save_params()

    os.remove(filename)
